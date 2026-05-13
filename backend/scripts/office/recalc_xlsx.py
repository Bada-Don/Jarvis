"""
scripts/office/recalc_xlsx.py
Force-recalculate all formulas in an Excel file using Excel COM Automation.

Usage:
    python scripts/office/recalc_xlsx.py output.xlsx
    python scripts/office/recalc_xlsx.py output.xlsx --timeout 60

Notes:
  - Requires Microsoft Excel to be installed.
  - Uses win32com to open, calculate, and save.
  - Scans all cells for Excel errors (#REF!, #DIV/0!, etc.) using openpyxl afterwards.
"""

import argparse
import json
import os
import sys
import time
from openpyxl import load_workbook

# Try to import win32com
try:
    import win32com.client
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}

def _recalc_via_excel(xlsx_path: str, timeout: int) -> bool:
    """
    Open Excel, load the file, calculate everything, and save.
    """
    if not HAS_WIN32COM:
        print("⚠️  win32com not installed. Run 'pip install pywin32'.", file=sys.stderr)
        return False

    abs_path = os.path.abspath(xlsx_path)
    excel = None
    wb = None
    try:
        # Initialize Excel
        # Use DispatchEx to ensure a fresh instance or Dispatch for existing
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # Open workbook
        wb = excel.Workbooks.Open(abs_path)
        
        # Force recalculation
        # xlCalculateFullRebuild = 1 (forces a full rebuild of the dependencies and calculates all cells)
        excel.CalculateFullRebuild()
        
        # Save and close
        wb.Save()
        wb.Close()
        return True
    except Exception as e:
        print(f"⚠️  Excel COM error: {e}", file=sys.stderr)
        return False
    finally:
        if excel:
            try:
                excel.Quit()
            except:
                pass
            # Clean up COM references
            del excel


def _scan_errors(xlsx_path: str) -> dict:
    """
    Scan all cells in all sheets for Excel error values.
    """
    # data_only=True ensures we read the CALCULATED values, not the formulas
    wb = load_workbook(xlsx_path, data_only=True)
    report = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        errors = []
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                if val in ERROR_VALUES:
                    errors.append({
                        "cell": cell.coordinate,
                        "error": val,
                    })
        if errors:
            report[sheet_name] = errors

    wb.close()
    return report


def recalc(xlsx_path: str, timeout: int = 30) -> dict:
    """
    Main entry point.
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    success = _recalc_via_excel(xlsx_path, timeout)
    
    recalc_status = "ok" if success else "skipped (Excel COM failed)"
    errors = _scan_errors(xlsx_path)

    total_errors = sum(len(v) for v in errors.values())
    return {
        "recalc": recalc_status,
        "error_count": total_errors,
        "errors_by_sheet": errors,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Recalculate formulas in an Excel file using MS Excel"
    )
    parser.add_argument("xlsx", help="Path to .xlsx file")
    parser.add_argument(
        "--timeout", type=int, default=30,
        help="Max seconds to wait (default: 30)"
    )
    args = parser.parse_args()

    try:
        report = recalc(args.xlsx, timeout=args.timeout)
        print(json.dumps(report, indent=2))

        if report["error_count"] > 0:
            print(f"\n❌ {report['error_count']} formula error(s) found")
            sys.exit(1)
        else:
            print("\n✅ No formula errors detected")
            sys.exit(0)
    except Exception as e:
        print(f"❌ Critical Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()

import os
import subprocess
import platform
import openpyxl
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from typing import Literal
from openai import OpenAI
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Initialize OpenAI client with API key from .env
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# --- 1. Structured Output Schema (Upgraded with Actions) ---
class ExcelCommand(BaseModel):
    action: Literal["edit_cell", "insert_row", "delete_row"]
    sheet_name: str
    target: str  # For edits: "A7". For inserts/deletes: "7"
    value: str   # The new value (leave empty for inserts/deletes)

class FileEdits(BaseModel):
    commands: list[ExcelCommand]

# --- 2. Type Inference Helper ---
def parse_excel_value(val_str: str):
    """Converts strings to int, float, boolean, or Formula."""
    val_str = val_str.strip()
    if not val_str: return ""
    if val_str.startswith("="): return val_str
    try: return int(val_str)
    except ValueError: pass
    try: return float(val_str)
    except ValueError: pass
    if val_str.lower() == "true": return True
    if val_str.lower() == "false": return False
    return val_str

# --- 3. Excel Helper Functions ---
def extract_excel_context(wb: openpyxl.Workbook, max_rows: int = 100) -> str:
    context = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        context.append(f"--- SHEET: {sheet_name} ---")
        
        max_col = ws.max_column
        col_headers = ["ROW/COL"] + [get_column_letter(i) for i in range(1, max_col + 1)]
        context.append(" | ".join(col_headers))
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx > max_rows:
                context.append(f"... (Truncated at {max_rows} rows) ...")
                break
            
            row_data = [str(row_idx)]
            for cell_val in row:
                val_str = str(cell_val) if cell_val is not None else ""
                val_str = val_str.replace("\n", " ")
                row_data.append(val_str)
                
            context.append(" | ".join(row_data))
        context.append("\n")
        
    return "\n".join(context)

def apply_excel_edits(wb: openpyxl.Workbook, commands: list[ExcelCommand]) -> list:
    """Executes structural changes (inserts/deletes) and edits sequentially."""
    diff_records = []
    
    for cmd in commands:
        if cmd.sheet_name not in wb.sheetnames:
            print(f"[WARNING] Sheet '{cmd.sheet_name}' does not exist.")
            continue
            
        ws = wb[cmd.sheet_name]
        
        try:
            if cmd.action == "insert_row":
                row_idx = int(cmd.target)
                ws.insert_rows(row_idx)
                diff_records.append({"action": "insert", "sheet": cmd.sheet_name, "target": cmd.target})
                
            elif cmd.action == "delete_row":
                row_idx = int(cmd.target)
                ws.delete_rows(row_idx)
                diff_records.append({"action": "delete", "sheet": cmd.sheet_name, "target": cmd.target})
                
            elif cmd.action == "edit_cell":
                target_cell = ws[cmd.target]
                old_value = str(target_cell.value) if target_cell.value is not None else "(empty)"
                
                parsed_value = parse_excel_value(cmd.value)
                target_cell.value = parsed_value
                
                diff_records.append({
                    "action": "edit",
                    "sheet": cmd.sheet_name,
                    "target": cmd.target,
                    "old": old_value,
                    "new": cmd.value,
                    "type": type(parsed_value).__name__
                })
        except Exception as e:
            print(f"[WARNING] Failed to execute {cmd.action} on '{cmd.target}': {e}")
            
    return diff_records

# --- 4. Core AI Function ---
def get_ai_edits(file_content: str, prompt: str) -> FileEdits:
    system_prompt = """
    You are an expert Data Analyst and Excel editor. 
    You have been provided with a text representation of an Excel workbook.
    
    CRITICAL RULES:
    1. ACTIONS: You can 'insert_row', 'delete_row', or 'edit_cell'. 
    2. SEQUENTIAL LOGIC: Commands are executed in the exact order you provide. If you use 'insert_row' at target '7', a new blank row appears at 7, and the old row 7 shifts down to 8. 
    3. TARGETING: For row operations, the 'target' is the row number (e.g., '7'). For cell edits, the target is the coordinate (e.g., 'A7').
    4. ADDING DATA: If asked to add a new row between existing data, first use 'insert_row', then follow up with multiple 'edit_cell' commands targeting that newly created blank row.
    5. DATA TYPES: Output just the numbers if it's a number. Output exact formulas starting with '=' if requested.
    """

    completion = client.beta.chat.completions.parse(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"Excel Data:\n```text\n{file_content}\n```\n\nUser Request: {prompt}"}
        ],
        response_format=FileEdits,
    )

    return completion.choices[0].message.parsed

# --- 5. Utilities ---
def show_diff(diff_records: list):
    if not diff_records:
        print("No changes proposed.")
        return
        
    print("\n--- PROPOSED EXCEL CHANGES ---")
    for rec in diff_records:
        if rec['action'] == 'insert':
            print(f"[{rec['sheet']}] ➕ Inserted new blank row at Row {rec['target']}")
        elif rec['action'] == 'delete':
            print(f"[{rec['sheet']}] ❌ Deleted Row {rec['target']}")
        elif rec['action'] == 'edit':
            print(f"[{rec['sheet']}] ✏️  {rec['target']}: '{rec['old']}' ---> '{rec['new']}' (Saved as {rec['type']})")
    print("------------------------------\n")

def open_file_in_editor(filepath: str):
    if platform.system() == 'Darwin': subprocess.call(('open', filepath))
    elif platform.system() == 'Windows': os.startfile(filepath)
    else: subprocess.call(('xdg-open', filepath))

# --- 6. Main Workflow ---
def main():
    print("🤖 AI Excel Spreadsheet Editor v3 (Structural Operations)")
    
    filepath = input("1. Enter the path to the .xlsx file: ").strip()
    if not os.path.isfile(filepath) or not filepath.endswith('.xlsx'):
        print(f"Error: Valid .xlsx file not found at '{filepath}'.")
        return

    wb_for_context = openpyxl.load_workbook(filepath, data_only=True)
    excel_text_grid = extract_excel_context(wb_for_context)
    wb_for_context.close()

    prompt = input('2. What would you like to do?: \n> ').strip()
    print("\nThinking and generating edits...")
    
    try:
        parsed_edits = get_ai_edits(excel_text_grid, prompt)
    except Exception as e:
        print(f"Failed to communicate with AI: {e}")
        return

    wb_for_editing = openpyxl.load_workbook(filepath)
    diff_records = apply_excel_edits(wb_for_editing, parsed_edits.commands)
    
    if not diff_records:
        print("No valid changes were applied. Exiting.")
        return

    show_diff(diff_records)

    confirm = input("4. Apply these changes and save? (y/n): ").strip().lower()
    if confirm == 'y':
        wb_for_editing.save(filepath)
        print(f"Success! Changes saved to {filepath}.")
        print("5. Opening Excel...")
        open_file_in_editor(filepath)
    else:
        print("Changes discarded.")

if __name__ == "__main__":
    main()
